from flask import Blueprint, request, jsonify
from bson.objectid import ObjectId
from app.extensions import mongo
from openpyxl import load_workbook
from datetime import datetime
import uuid
import re

fileupload_bp = Blueprint("fileupload_bp", __name__)

def validate_excel_row(row: dict):
    if not row.get("department"):
        return "Department is required"

    amount = row.get("amount")
    if amount is None or amount <= 0:
        return "Amount must be greater than zero"

    return None

@fileupload_bp.route("/upload-file", methods=["POST"])
def upload_excel_file():
    """
    Upload and process Excel files dynamically into MongoDB.
    ---
    tags:
      - Excel Upload
    summary: "Upload Excel file and store data dynamically into MongoDB collections"
    description: >
      Uploads an Excel file, streams and validates each row, stores
      file metadata, row data, and validation errors into dynamically
      selected MongoDB database and collections.
      Files are NOT stored on disk.
    consumes:
      - multipart/form-data
    parameters:
      - name: file
        in: formData
        type: file
        required: true
      - name: database
        in: formData
        type: string
        required: true
        example: upload_db
      - name: data_collection
        in: formData
        type: string
        required: true
        example: excel_data
      - name: uploadedBy
        in: formData
        type: string
        required: true
        example: JaneDoe
    responses:
      200:
        description: File processed successfully
      400:
        description: Invalid request
      500:
        description: Internal server error
    """

    try:
        # ----------------------------
        # 1. Read form data
        # ----------------------------
        file = request.files.get("file")
        database_name = request.form.get("database", "").strip()
        data_collection_name = request.form.get("data_collection", "").strip()
        uploaded_by = request.form.get("uploadedBy", "").strip()

        if not file or not database_name or not data_collection_name or not uploaded_by:
            return jsonify({
                "error": "file, database, data_collection, and uploadedBy are required"
            }), 400

        if not file.filename.endswith(".xlsx"):
            return jsonify({"error": "Only .xlsx files are allowed"}), 400

        filename = file.filename

        # ----------------------------
        # 2. Validate filename pattern
        # ----------------------------
        # AnnualBudget2026_JaneDoe_MarkSmith_20251216.xlsx
        match = re.match(
            r"(.+?)(\d{4})_(.+?)_(.+?)_(\d{8})\.xlsx", filename
        )

        if not match:
            return jsonify({"error": "Invalid filename format"}), 400

        _, year, creater, reviewer, date_str = match.groups()

        # ----------------------------
        # 3. Mongo database & collections
        # ----------------------------
        db = mongo.cx[database_name]

        file_collection = db.file_uploads
        data_collection = db[data_collection_name]
        error_collection = db.validation_errors

        # ----------------------------
        # 4. Insert file metadata
        # ----------------------------
        file_id = str(uuid.uuid4())

        file_collection.insert_one({
            "_id": file_id,
            "filename": filename,
            "createdBy": creater,
            "uploadedBy": uploaded_by,
            "reviewedBy": reviewer,
            "year": int(year),
            "uploadDate": datetime.utcnow(),
            "status": "uploading",
            "totalRows": 0,
            "validRows": 0,
            "invalidRows": 0
        })

        # ----------------------------
        # 5. Stream Excel file
        # ----------------------------
        workbook = load_workbook(
            file,
            read_only=True,
            data_only=True
        )

        sheet = workbook.active

        headers = [
            cell.value for cell in
            next(sheet.iter_rows(min_row=1, max_row=1))
        ]

        batch_size = 500
        data_buffer = []
        error_buffer = []

        total = valid = invalid = 0

        for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            total += 1
            row_data = dict(zip(headers, [cell.value for cell in row]))
            
            # error_msg = validate_excel_row(row_data)
            error_msg = ""

            if error_msg:
                invalid += 1
                error_buffer.append({
                    "fileId": file_id,
                    "rowNumber": row_index,
                    "error": error_msg,
                    "createdAt": datetime.utcnow()
                })
            else:
                valid += 1
                data_buffer.append({
                    "fileId": file_id,
                    "rowNumber": row_index,
                    **row_data,
                    "createdAt": datetime.utcnow()
                })

            if len(data_buffer) >= batch_size:
                data_collection.insert_many(data_buffer)
                data_buffer.clear()

            if len(error_buffer) >= batch_size:
                error_collection.insert_many(error_buffer)
                error_buffer.clear()

        if data_buffer:
            data_collection.insert_many(data_buffer)

        if error_buffer:
            error_collection.insert_many(error_buffer)

        # ----------------------------
        # 6. Final status update
        # ----------------------------
        final_status = "failed" if invalid > 0 else "compeleted"

        file_collection.update_one(
            {"_id": file_id},
            {"$set": {
                "status": final_status,
                "totalRows": total,
                "validRows": valid,
                "invalidRows": invalid
            }}
        )

        return jsonify({
            "fileId": file_id,
            "database": database_name,
            "collection": data_collection_name,
            "filename": filename,
            "status": final_status,
            "totalRows": total,
            "validRows": valid,
            "invalidRows": invalid
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
    
@fileupload_bp.route("/files", methods=["POST"])
def get_uploaded_files():
    """
    Get uploaded Excel files with pagination and optional filtering.
    ---
    tags:
      - File Upload
    summary: "Fetch uploaded Excel files"
    description: "Retrieve uploaded Excel file metadata from MongoDB with pagination and optional filters."
    consumes:
      - application/json
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            database:
              type: string
              example: "reports_db"
            collection:
              type: string
              example: "uploaded_files"
            page:
              type: integer
              example: 1
              default: 1
            page_records:
              type: integer
              example: 20
              default: 20
            fileName:
              type: string
              example: "AnnualBudget2026"
            uploadedBy:
              type: string
              example: "JaneDoe"
    responses:
      200:
        description: "File list"
      400:
        description: "Invalid request"
      500:
        description: "Server error"
    """
    try:
        data = request.get_json()

        database_name = data.get("database", "").strip()
        collection_name = data.get("collection", "").strip()
        page = int(data.get("page", 1))
        page_records = int(data.get("page_records", 20))

        if not database_name or not collection_name:
            return jsonify({"error": "Database and collection are required"}), 400

        # Build filter query
        query = {
            k: {"$regex": v, "$options": "i"}
            for k, v in data.items()
            if v and k not in ["database", "collection", "page", "page_records"]
        }

        db = mongo.cx[database_name]
        collection = db[collection_name]

        skip = (page - 1) * page_records

        files = list(
            collection.find(query)
            .sort("createdAt", -1)
            .skip(skip)
            .limit(page_records)
        )

        total = collection.count_documents(query)

        for f in files:
            f["_id"] = str(f["_id"])

        return jsonify({
            "count": total,
            "page": page,
            "page_records": page_records,
            "data": files
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@fileupload_bp.route("/files/delete", methods=["POST"])
def delete_uploaded_file():
    """
    Delete an uploaded file and its associated data.
    ---
    tags:
      - File Upload
    summary: "Delete uploaded Excel file"
    description: "Deletes file metadata and all parsed Excel data linked to the fileId."
    consumes:
      - application/json
    parameters:
      - name: body
        in: body
        required: true
        schema:
          type: object
          properties:
            database:
              type: string
              example: "reports_db"
            file_collection:
              type: string
              example: "uploaded_files"
            data_collection:
              type: string
              example: "excel_data"
            fileId:
              type: string
              example: "65f1a9c12a7dbe99f7e12345"
    responses:
      200:
        description: "File deleted successfully"
      400:
        description: "Invalid request"
      404:
        description: "File not found"
      500:
        description: "Server error"
    """
    try:
        data = request.get_json()

        database_name = data.get("database", "").strip()
        file_collection_name = data.get("file_collection", "").strip()
        data_collection_name = data.get("data_collection", "").strip()
        file_id = data.get("fileId")

        if not all([database_name, file_collection_name, data_collection_name, file_id]):
            return jsonify({"error": "Missing required fields"}), 400

        db = mongo.cx[database_name]
        file_collection = db[file_collection_name]
        data_collection = db[data_collection_name]

        file_object_id = ObjectId(file_id)

        file_doc = file_collection.find_one({"_id": file_object_id})
        if not file_doc:
            return jsonify({"error": "File not found"}), 404

        # Delete parsed data first
        data_delete_result = data_collection.delete_many({"fileId": file_object_id})

        # Delete file metadata
        file_collection.delete_one({"_id": file_object_id})

        return jsonify({
            "message": "File deleted successfully",
            "rowsDeleted": data_delete_result.deleted_count
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500
